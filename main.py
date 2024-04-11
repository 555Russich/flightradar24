import logging
import random
import time
from pathlib import Path
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import tqdm

from my_logging import get_logger

LOGIN = 'mr.unlife@gmail.com'
PASSWORD = 'aweryFL24'
DATE_FROM = '01.01.2023'

FILEPATH_LOGGER = Path('flightradar.log')
FILEPATH_INPUT_TXT = Path('input.txt')
FILEPATH_AIRPORTS = Path('input_airports.txt')
FILEPATH_OUTPUT_XLSX = Path('output.xlsx')

MAX_RETRIES = 5
RETRY_SLEEP_RANGE = (30, 60)
DEFAULT_SLEEP = 1
DOMAIN = 'https://www.flightradar24.com'
HEADERS = {
    'authority': 'api.flightradar24.com',
    'accept': '*/*',
    'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'origin': 'https://www.flightradar24.com',
    'referer': 'https://www.flightradar24.com/',
    'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-site',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
}


class Scraper:
    def __init__(self):
        self._session = requests.Session()

    @property
    def token(self) -> str:
        return self._session.cookies.get('_frPl')

    def login(self, login: str, password: str):
        headers = HEADERS.copy()
        headers['content-type'] = 'multipart/form-data; boundary=----WebKitFormBoundarywznYKz3LiItMDEVe'
        data = f'''------WebKitFormBoundarywznYKz3LiItMDEVe
Content-Disposition: form-data; name="email"

{login}
------WebKitFormBoundarywznYKz3LiItMDEVe
Content-Disposition: form-data; name="password"

{password}
------WebKitFormBoundarywznYKz3LiItMDEVe
Content-Disposition: form-data; name="remember"

true
------WebKitFormBoundarywznYKz3LiItMDEVe
Content-Disposition: form-data; name="type"

web
------WebKitFormBoundarywznYKz3LiItMDEVe--
'''
        resp = self._session.post(url=f'{DOMAIN}/user/login', headers=headers, data=data)
        if resp.status_code != 200:
            raise Exception(f'Logging failed {resp.status_code=}')

    def _get_request(
            self,
            url: str,
            retries: int = MAX_RETRIES,
            sleep_range: tuple[int, int] = RETRY_SLEEP_RANGE,
            **kwargs
    ) -> requests.Response:
        for retry in range(1, retries+1):
            response = requests.get(url, **kwargs)
            if response.ok:
                return response

            sleep_time = random.randint(*sleep_range)
            logging.warning(f'Retry {retry}/{retries} | {response.status_code=}! Sleeping for {sleep_time} seconds')
            time.sleep(sleep_time)
        raise Exception(f'Failed {retries} retries!')

    def get_existing_airlines(self) -> list[tuple[str, str]]:
        response = self._get_request(f'{DOMAIN}/data/airlines', headers=HEADERS)
        soup = BeautifulSoup(response.text, 'html5lib')
        table = soup.find('table', {'id': 'tbl-datatable'})

        existing_airlines = []
        for td in table.find_all('td', class_='notranslate'):
            a = td.find('a')
            text = a.get_text()
            href = a.get('href')
            existing_airlines.append((text, href))
        return existing_airlines

    def get_airline_fleet(self, airline_href: str) -> list[str]:
        response = self._get_request(f'{DOMAIN}{airline_href}/fleet', headers=HEADERS)
        soup = BeautifulSoup(response.text, 'html5lib')
        return [a.text.strip().lower() for a in soup.find_all('a', class_='regLinks')]

    def get_aircraft_history(self, aircraft_number: str, earliest_date: datetime | None = None) -> list[dict]:
        headers = HEADERS.copy()
        params = {
            'query': aircraft_number.lower(),
            'fetchBy': 'reg',
            'page': 1,
            'pk': '',
            'limit': '100',
            'token': self.token,
            'timestamp': '',
            'olderThenFlightId': '',
        }
        history = []

        while True:
            logging.info(f'Sending request to page={params["page"]} aircraft={aircraft_number}')
            response = self._get_request(
                url='https://api.flightradar24.com/common/v1/flight/list.json',
                headers=headers,
                params=params
            )
            data = response.json()['result']['response']
            aircraft = data['aircraftInfo']

            if data['data'] is None:
                return []

            for d in data.get('data', []):
                text_status = d['status']['text']
                if text_status in ['Scheduled']:
                    continue

                departure_timestamp = d['time']['scheduled']['departure']
                updated_timestamp = d['time']['other']['updated']
                if departure_timestamp:
                    date = datetime.fromtimestamp(departure_timestamp)
                else:
                    date = datetime.fromtimestamp(updated_timestamp)

                if earliest_date and date < earliest_date:
                    return history

                duration = d['time']['other']['duration']

                if d['airport']['origin']:
                    from_city = d['airport']['origin']['position']['region']['city']
                    from_iata = d['airport']['origin']['code']['iata']
                    from_ = f'{from_city} ({from_iata})'
                else:
                    from_ = None

                if d['airport']['destination']:
                    to_city = d['airport']['destination']['position']['region']['city']
                    to_iata = d['airport']['destination']['code']['iata']
                    to = f'{to_city} ({to_iata})'
                else:
                    to = None

                flight = {
                    'NUMBER': aircraft_number.upper(),
                    'AIRLINE': aircraft['airline']['name'],
                    'MODEL': aircraft['model']['text'],
                    'DATE': date.strftime('%d.%m.%Y'),
                    'FROM': from_,
                    'TO': to,
                    'FLIGHT': d['identification']['number']['default'],
                    'FLIGHT TIME': time.strftime('%H:%M', time.gmtime(duration)) if duration else None,
                    'STATUS': text_status
                }
                if flight in history:
                    logging.debug(f'Found duplicate {flight=}')
                else:
                    history.append(flight)

            if data['page']['more']:
                params['page'] += 1
                params['olderThenFlightId'] = data['data'][-1]['identification']['id']
                last_dept_timestamp = data['data'][-1]['time']['scheduled']['departure']
                if last_dept_timestamp:
                    params['timestamp'] = last_dept_timestamp
                time.sleep(3)
            else:
                return history

    def get_airport_history(self, airport_code: str, earliest_date: datetime | None = None) -> list[dict]:
        return (self._get_airport_history(airport_code=airport_code, direction='arrivals', earliest_date=earliest_date) +
                self._get_airport_history(airport_code=airport_code, direction='departures', earliest_date=earliest_date))

    def _get_airport_history(self, airport_code: str, direction: str, earliest_date: datetime | None = None) -> list[dict]:
        page = 1
        history = []

        while True:
            params = {
                'code': airport_code.lower(),
                'plugin[]': '',
                'plugin-setting[schedule][mode]': direction,
                'plugin-setting[schedule][timestamp]': int((datetime.now() - timedelta(minutes=15)).timestamp()),
                'page': page,
                'limit': '100',
                'fleet': '',
                'token': self.token,
            }
            logging.info(f'Sending request to page={params["page"]} airport={airport_code}')
            response = self._get_request(
                url='https://api.flightradar24.com/common/v1/airport.json',
                headers=HEADERS,
                params=params
            )
            r_json = response.json()['result']['response']
            data = r_json['airport']['pluginData']['schedule'][direction]

            for d in data['data']:
                d = d['flight']
                duration = d['time']['other']['duration']
                date = datetime.fromtimestamp(d['time']['scheduled']['departure'])

                if d['status']['text'] == 'Scheduled' or d['status']['text'].startswith('Estimated'):
                    continue

                if earliest_date and date < earliest_date:
                    return history

                origin, destination = self._parse_origin_and_destination(
                    d=d, airport=r_json['airport'], direction=direction)

                flight = {
                    'NUMBER': d['aircraft']['registration'],
                    'AIRLINE': d['airline']['name'] if d['airline'] else None,
                    'MODEL': d['aircraft']['model']['text'],
                    'DATE': date.strftime('%d.%m.%Y'),
                    'FROM': origin,
                    'TO': destination,
                    'FLIGHT': d['identification']['number']['default'],
                    'FLIGHT TIME': time.strftime('%H:%M', time.gmtime(duration)) if duration else None,
                    'STATUS': d['status']['text']
                }

                if flight not in history:
                    history.append(flight)

            if data['page']['current'] == data['page']['total']:
                return history
            page -= 1

    def _parse_origin_and_destination(self, d: dict, airport: dict, direction: str) -> tuple[str, str]:
        airport_name = airport['pluginData']['details']['name']
        airport_iata = airport['pluginData']['details']['code']['iata']
        airport_icao = airport['pluginData']['details']['code']['icao']

        if direction == 'arrivals':
            if d['airport']['origin']:
                from_city = d['airport']['origin']['position']['region']['city']
                from_iata = d['airport']['origin']['code']['iata']
                origin = f'{from_city} ({from_iata})'
            else:
                origin = None
            destination = f'{airport_name} ({airport_iata}/{airport_icao})'
        elif direction == 'departures':
            origin = f'{airport_name} ({airport_iata}/{airport_icao})'
            if d['airport']['destination']:
                from_city = d['airport']['destination']['position']['region']['city']
                from_iata = d['airport']['destination']['code']['iata']
                destination = f'{from_city} ({from_iata})'
            else:
                destination = None
        else:
            raise Exception(f'{direction=}')

        return origin, destination


def define_inputs_type(existing_airlines: list[tuple[str, str]], inputs: list[str]) -> dict[str, list]:
    defined_inputs = {
        'aircrafts': [],
        'airlines': [],
    }

    for line in inputs:
        airline = None
        for ex_airline in existing_airlines:
            if ex_airline[0].lower() == line:
                airline = ex_airline
                break

        if airline:
            defined_inputs['airlines'].append(airline)
        else:
            defined_inputs['aircrafts'].append(line)
    return defined_inputs


def get_raw_inputs(filepath: Path) -> list[str]:
    if filepath.exists() is False:
        raise FileNotFoundError(f'Need to create {filepath}')

    return [x.strip().lower() for x in filepath.read_text(encoding='utf-8').split('\n') if x]


def read_from_excel(filepath: Path) -> list[dict]:
    wb = load_workbook(filepath)
    rows = list(wb.active.iter_rows(values_only=True))
    return [{k: v for k, v in zip(rows[0], row)} for row in rows[1:]]


def write_to_excel(filepath: Path, data: list[dict]) -> None:
    if filepath.exists():
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        columns = list(data[0].keys())
        ws.append(columns)

    for d in data:
        ws.append(list(d.values()))

    wb.save(filepath)


def main():
    raw_inputs = get_raw_inputs(FILEPATH_INPUT_TXT)
    date_from = datetime.strptime(DATE_FROM, '%d.%m.%Y') if DATE_FROM else None

    scraper = Scraper()
    if LOGIN and PASSWORD:
        scraper.login(login=LOGIN, password=PASSWORD)
        logging.info(f'Logged in successfully')

    existing_airlines = scraper.get_existing_airlines()
    defined_inputs = define_inputs_type(existing_airlines=existing_airlines, inputs=raw_inputs)

    aircrafts_to_scrap = defined_inputs['aircrafts']
    for airline_data in defined_inputs['airlines']:
        airline_fleet = scraper.get_airline_fleet(airline_data[1])
        aircrafts_to_scrap += airline_fleet
        logging.info(f'Got {len(airline_fleet)} aircrafts for "{airline_data[0]}" airline')
    airports_to_scrap = [s.replace('\n', '').lower() for s in FILEPATH_AIRPORTS.read_text().split('\n')]
    logging.info(f'\naircrafts={defined_inputs["aircrafts"]}\nairlines={defined_inputs["airlines"]}'
                 f'\nairports={airports_to_scrap}')

    if FILEPATH_OUTPUT_XLSX.exists():
        data_in_xlsx = read_from_excel(FILEPATH_OUTPUT_XLSX)
    else:
        data_in_xlsx = []

    scrapped_data = []
    data_to_append = []

    for airport_code in airports_to_scrap:
        airport_flights = scraper.get_airport_history(airport_code=airport_code, earliest_date=date_from)
        scrapped_data += airport_flights
        time.sleep(DEFAULT_SLEEP)

    for aircraft_number in tqdm.tqdm(aircrafts_to_scrap):
        aircraft_flights = scraper.get_aircraft_history(aircraft_number, earliest_date=date_from)
        scrapped_data += aircraft_flights
        time.sleep(DEFAULT_SLEEP)

    for flight in scrapped_data:
        if flight not in data_in_xlsx:
            data_to_append.append(flight)

    logging.info(f'Appending {len(data_to_append)} new flights to "{FILEPATH_OUTPUT_XLSX.name}"')
    write_to_excel(FILEPATH_OUTPUT_XLSX, data_to_append)


if __name__ == '__main__':
    get_logger(FILEPATH_LOGGER)
    main()
